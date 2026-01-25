#!/usr/bin/env python3
"""
Servidor simples para receber e armazenar votos dos itens Luna Obscura
Salva os votos em CSV e XLSX

Para rodar:
    python server_votes.py

O servidor ficará rodando em http://localhost:5000
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import csv
import os
from datetime import datetime
import json

# Tenta importar openpyxl para XLSX, se não tiver, só usa CSV
try:
    from openpyxl import Workbook, load_workbook
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False
    print("⚠️  openpyxl não instalado. Instale com: pip install openpyxl")
    print("   Por enquanto, apenas CSV será usado.")

app = Flask(__name__)
CORS(app)  # Permite requisições de qualquer origem

# Arquivos de dados
CSV_FILE = 'luna_votes.csv'
XLSX_FILE = 'luna_votes.xlsx'
JSON_FILE = 'luna_votes_summary.json'

def init_files():
    """Inicializa os arquivos se não existirem"""
    # CSV
    if not os.path.exists(CSV_FILE):
        with open(CSV_FILE, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['timestamp', 'item_name', 'vote_type', 'action'])
    
    # JSON summary
    if not os.path.exists(JSON_FILE):
        with open(JSON_FILE, 'w', encoding='utf-8') as f:
            json.dump({}, f)
    
    # XLSX
    if HAS_XLSX and not os.path.exists(XLSX_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Votos"
        ws.append(['Timestamp', 'Item', 'Tipo de Voto', 'Ação'])
        wb.save(XLSX_FILE)

def save_vote_to_csv(item_name, vote_type, action):
    """Salva voto no CSV"""
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    with open(CSV_FILE, 'a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([timestamp, item_name, vote_type, action])

def save_vote_to_xlsx(item_name, vote_type, action):
    """Salva voto no XLSX"""
    if not HAS_XLSX:
        return
    
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    wb = load_workbook(XLSX_FILE)
    ws = wb.active
    ws.append([timestamp, item_name, vote_type, action])
    wb.save(XLSX_FILE)

def update_summary(item_name, vote_type, action):
    """Atualiza o resumo de votos em JSON"""
    # Lê resumo atual
    if os.path.exists(JSON_FILE):
        with open(JSON_FILE, 'r', encoding='utf-8') as f:
            summary = json.load(f)
    else:
        summary = {}
    
    # Inicializa item se não existir
    if item_name not in summary:
        summary[item_name] = {'positive': 0, 'neutral': 0, 'negative': 0}
    
    # Atualiza contadores
    if action == 'add':
        summary[item_name][vote_type] = summary[item_name].get(vote_type, 0) + 1
    elif action == 'remove':
        summary[item_name][vote_type] = max(0, summary[item_name].get(vote_type, 0) - 1)
    elif action == 'change':
        # Remove voto anterior e adiciona novo
        # Isso é tratado por duas chamadas separadas (remove + add)
        pass
    
    # Salva resumo
    with open(JSON_FILE, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)
    
    return summary

@app.route('/api/vote', methods=['POST'])
def vote():
    """Endpoint para receber votos"""
    try:
        data = request.json
        item_name = data.get('item_name')
        vote_type = data.get('vote_type')  # 'positive', 'neutral', 'negative'
        action = data.get('action')  # 'add', 'remove', 'change'
        
        if not item_name or not vote_type or not action:
            return jsonify({'error': 'Dados incompletos'}), 400
        
        # Salva nos arquivos
        save_vote_to_csv(item_name, vote_type, action)
        save_vote_to_xlsx(item_name, vote_type, action)
        summary = update_summary(item_name, vote_type, action)
        
        return jsonify({
            'success': True,
            'summary': summary.get(item_name, {'positive': 0, 'neutral': 0, 'negative': 0})
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/summary', methods=['GET'])
def get_summary():
    """Endpoint para obter resumo de todos os votos"""
    try:
        if os.path.exists(JSON_FILE):
            with open(JSON_FILE, 'r', encoding='utf-8') as f:
                summary = json.load(f)
        else:
            summary = {}
        
        return jsonify({'success': True, 'summary': summary})
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/health', methods=['GET'])
def health():
    """Endpoint de health check"""
    return jsonify({'status': 'ok', 'has_xlsx': HAS_XLSX})

if __name__ == '__main__':
    print("🚀 Iniciando servidor de votos...")
    init_files()
    print(f"📁 Arquivos de dados:")
    print(f"   - CSV: {CSV_FILE}")
    if HAS_XLSX:
        print(f"   - XLSX: {XLSX_FILE}")
    print(f"   - JSON: {JSON_FILE}")
    
    # Porta do ambiente (para produção) ou 5000 (desenvolvimento)
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    print(f"\n🌐 Servidor rodando em: http://0.0.0.0:{port}")
    print(f"📊 Endpoints disponíveis:")
    print(f"   - POST /api/vote - Enviar voto")
    print(f"   - GET /api/summary - Obter resumo de votos")
    print(f"   - GET /api/health - Health check")
    print(f"\n⚠️  Para parar o servidor, pressione Ctrl+C\n")
    
    app.run(host='0.0.0.0', port=port, debug=debug)
